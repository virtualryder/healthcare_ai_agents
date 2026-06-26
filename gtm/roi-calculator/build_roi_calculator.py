#!/usr/bin/env python3
"""Generate the SA-fillable HPP ROI Calculator workbook (live Excel formulas).
Inputs are editable (orange); outputs compute automatically (no hard-coded results).
Every benchmark default traces to ../HPP-DECK-SOURCES.md. Run: python build_roi_calculator.py"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INK="FF232F3E"; ORANGE="FFFF9900"; TEAL="FF01A88D"; LIGHT="FFF4F7FA"; WHITE="FFFFFFFF"
hdr_fill=PatternFill("solid", fgColor=INK); in_fill=PatternFill("solid", fgColor="FFFFF3E0")
out_fill=PatternFill("solid", fgColor="FFE7F5F2"); band=PatternFill("solid", fgColor=LIGHT)
H=Font(bold=True, color=WHITE, size=13, name="Calibri")
LBL=Font(color="FF232F3E", size=11, name="Calibri"); BOLD=Font(bold=True, size=11, name="Calibri")
thin=Side(style="thin", color="FFD6DEE6"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook()

def title_row(ws, text, span=3):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1,text); c.fill=hdr_fill; c.font=H; c.alignment=Alignment(vertical="center")
    ws.row_dimensions[1].height=26

def section(ws, r, text, span=3):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=span)
    c=ws.cell(r,1,text); c.font=BOLD; c.fill=band

def inp(ws, r, label, value, fmt="0.00"):
    ws.cell(r,1,label).font=LBL
    c=ws.cell(r,2,value); c.fill=in_fill; c.border=border; c.number_format=fmt; c.font=BOLD
def out(ws, r, label, formula, fmt='#,##0'):
    ws.cell(r,1,label).font=LBL
    c=ws.cell(r,2,formula); c.fill=out_fill; c.border=border; c.number_format=fmt; c.font=BOLD

# ── README ────────────────────────────────────────────────────────────────────
ws=wb.active; ws.title="README"; ws.column_dimensions['A'].width=110
title_row(ws,"HPP AI Agent Suite — ROI Calculator", span=1)
notes=[
 "", "How to use: edit the ORANGE input cells on each model tab; the TEAL output cells recompute automatically.",
 "Tabs: 'Revenue Cycle (01)', 'Prior Auth (02)', 'Capacity (generic)', and 'Suite Summary'.",
 "Every default is a benchmark from gtm/HPP-DECK-SOURCES.md (source-class tagged) — replace with the customer's actuals.",
 "",
 "What this is NOT: a guarantee. It is a decision-support model. The agents do not auto-submit, deny, or recoup;",
 "value comes from compressing analyst/clinician minutes and reducing leakage — measured in a pilot, not assumed.",
 "",
 "Maturity: the suite is Demonstrated + Deployable-by-design. Production-readiness (CSV/CSA, IdP, live-connector",
 "validation, penetration test, HITRUST) is the engagement.",
 "",
 "Benchmarks (defaults): initial denial rate ~11.8% and ~$18B/yr overturning denials [industry-research];",
 "~39 prior auths/physician/week (~13 hrs) [association]; healthcare call ~$25-$35 [industry-research].",
]
for i,t in enumerate(notes, start=2):
    ws.cell(i,1,t).font=Font(size=11, name="Calibri", bold=(t.startswith(("How to use","Tabs","What this is NOT","Maturity"))))

# ── Revenue Cycle (Agent 01) ──────────────────────────────────────────────────
ws=wb.create_sheet("Revenue Cycle (01)")
ws.column_dimensions['A'].width=52; ws.column_dimensions['B'].width=18
title_row(ws,"Agent 01 — Revenue-Cycle & Denial ROI")
section(ws,3,"Inputs (edit the orange cells)")
inp(ws,4,"Annual claims volume", 1200000, "#,##0")
inp(ws,5,"Initial denial rate", 0.118, "0.0%")
inp(ws,6,"Avg net revenue per denied claim ($)", 180, '$#,##0')
inp(ws,7,"Share of denials currently never reworked", 0.45, "0%")
inp(ws,8,"Recovery (overturn) lift on newly-worked denials", 0.55, "0%")
inp(ws,9,"Fully-loaded cost to rework one claim ($)", 57, '$#,##0')
inp(ws,10,"Agent deflection of manual rework effort", 0.50, "0%")
section(ws,12,"Outputs (compute automatically)")
out(ws,13,"Denied claims / year", "=B4*B5")
out(ws,14,"Recovered revenue (previously abandoned) / year", "=B13*B7*B8*B6", '$#,##0')
out(ws,15,"Rework cost avoided / year", "=B13*(1-B7)*B9*B10", '$#,##0')
out(ws,16,"Total modeled annual value", "=B14+B15", '$#,##0')

# ── Prior Auth (Agent 02) ─────────────────────────────────────────────────────
ws=wb.create_sheet("Prior Auth (02)")
ws.column_dimensions['A'].width=52; ws.column_dimensions['B'].width=18
title_row(ws,"Agent 02 — Prior-Authorization ROI")
section(ws,3,"Inputs (edit the orange cells)")
inp(ws,4,"Prior auths per provider per week", 39, "0")
inp(ws,5,"Providers in scope", 200, "#,##0")
inp(ws,6,"Staff minutes per prior auth (assembly/admin)", 20, "0")
inp(ws,7,"Agent deflection of assembly/admin time", 0.55, "0%")
inp(ws,8,"Loaded staff cost per hour ($)", 55, '$#,##0')
inp(ws,9,"Working weeks per year", 48, "0")
section(ws,11,"Outputs (compute automatically)")
out(ws,12,"Staff hours returned / year", "=B4*B5*(B6/60)*B7*B9", '#,##0')
out(ws,13,"Annualized cost avoided / year", "=B12*B8", '$#,##0')

# ── Capacity (generic, agents 03/05/06/07/08) ────────────────────────────────
ws=wb.create_sheet("Capacity (generic)")
ws.column_dimensions['A'].width=52; ws.column_dimensions['B'].width=18
title_row(ws,"Generic Capacity Model (03 / 05 / 06 / 07 / 08)")
section(ws,3,"Inputs (edit the orange cells)")
inp(ws,4,"Work items / month (visits, reviews, claims, calls)", 50000, "#,##0")
inp(ws,5,"Staff minutes per item today", 12, "0")
inp(ws,6,"Agent deflection / assist factor", 0.45, "0%")
inp(ws,7,"Loaded staff cost per hour ($)", 60, '$#,##0')
section(ws,9,"Outputs (compute automatically)")
out(ws,10,"Staff hours returned / month", "=B4*(B5/60)*B6", '#,##0')
out(ws,11,"Annualized capacity recovered ($)", "=B10*12*B7", '$#,##0')

# ── Suite Summary ─────────────────────────────────────────────────────────────
ws=wb.create_sheet("Suite Summary")
ws.column_dimensions['A'].width=52; ws.column_dimensions['B'].width=20
title_row(ws,"Suite Summary — modeled annual value")
ws.cell(3,1,"Agent 01 — Revenue Cycle & Denial").font=LBL
ws.cell(3,2,"='Revenue Cycle (01)'!B16").number_format='$#,##0'; ws.cell(3,2).fill=out_fill; ws.cell(3,2).font=BOLD
ws.cell(4,1,"Agent 02 — Prior-Authorization").font=LBL
ws.cell(4,2,"='Prior Auth (02)'!B13").number_format='$#,##0'; ws.cell(4,2).fill=out_fill; ws.cell(4,2).font=BOLD
ws.cell(5,1,"Agents 03/05/06/07/08 — Capacity (generic, per agent)").font=LBL
ws.cell(5,2,"='Capacity (generic)'!B11").number_format='$#,##0'; ws.cell(5,2).fill=out_fill; ws.cell(5,2).font=BOLD
ws.cell(7,1,"Indicative total (01 + 02 + one capacity agent)").font=BOLD
ws.cell(7,2,"=B3+B4+B5").number_format='$#,##0'; ws.cell(7,2).fill=PatternFill("solid",fgColor=ORANGE); ws.cell(7,2).font=Font(bold=True,size=12)
ws.cell(9,1,"Note: indicative only; scope the capacity model to each in-scope agent. Not a guarantee.").font=Font(italic=True,size=10)

wb.save("HPP-ROI-Calculator.xlsx")
print("wrote HPP-ROI-Calculator.xlsx with", len(wb.sheetnames), "sheets:", wb.sheetnames)
